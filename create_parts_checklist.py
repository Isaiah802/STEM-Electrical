"""
NLC STEM Club UAV - Parts List Checklist Generator
Creates comprehensive Excel checklist with all components
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_parts_checklist():
    """Generate comprehensive parts checklist Excel file"""
    
    print("=" * 80)
    print("NLC STEM Club UAV - Parts List Checklist Generator")
    print("=" * 80)
    print()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "UAV Parts Checklist"
    
    # Define styles
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(name='Helvetica', size=12, bold=True, color="FFFFFF")
    
    category_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    category_font = Font(name='Helvetica', size=11, bold=True, color="FFFFFF")
    
    subcategory_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    subcategory_font = Font(name='Helvetica', size=10, bold=True)
    
    normal_font = Font(name='Helvetica', size=10)
    bold_font = Font(name='Helvetica', size=10, bold=True)
    price_font = Font(name='Helvetica', size=10, bold=True, color="006100")
    
    link_font = Font(name='Helvetica', size=9, underline='single', color="0563C1")
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 5   # ✓ checkbox
    ws.column_dimensions['B'].width = 22  # Category
    ws.column_dimensions['C'].width = 32  # Component
    ws.column_dimensions['D'].width = 8   # Qty
    ws.column_dimensions['E'].width = 40  # Specifications
    ws.column_dimensions['F'].width = 10  # Status
    ws.column_dimensions['G'].width = 12  # Price
    ws.column_dimensions['H'].width = 30  # Notes/Links
    
    # Title Row
    ws.merge_cells('A1:H1')
    ws['A1'] = "NLC STEM CLUB - 3D-PRINTED SOLAR FLOATPLANE UAV"
    ws['A1'].font = Font(name='Helvetica', size=16, bold=True, color="003366")
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 25
    
    # Subtitle Row
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Complete Parts List & Procurement Checklist - Generated {datetime.now().strftime('%B %d, %Y')}"
    ws['A2'].font = Font(name='Helvetica', size=11, italic=True)
    ws['A2'].alignment = center_align
    ws.row_dimensions[2].height = 20
    
    # Header Row
    row = 4
    headers = ['✓', 'Category', 'Component / Description', 'Qty', 'Specifications', 'Status', 'Price (USD)', 'Notes / Links']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[row].height = 30
    
    # Parts data organized by category
    parts_data = [
        # AIRFRAME & STRUCTURE
        ("AIRFRAME & STRUCTURE", "", "", "", "", "", "", ""),
        ("", "Filament", "eSUN PLA LW PLA 3D Printer Filament", "1-2kg", "Lightweight PLA, Low Density, 1.75mm", "Ordered", "$45", "https://a.co/d/2AS4Lgc"),
        ("", "Reinforcement", "Carbon Fiber Rods - 2mm Wing Spars", "10pcs", "2mm diameter, 500mm length, Carbon Fiber", "Ordered", "$12", "Essential structural reinforcement"),
        ("", "Adhesives", "Foam Cure - Foam Safe CA Glue", "2 bottles", "Cyanoacrylate, Foam-safe formulation", "Ordered", "$15", "Included in Tools/Hardware"),
        ("", "Adhesives", "5-Minute Epoxy", "2 tubes", "Two-part epoxy for structural bonds", "Ordered", "$12", "Included in Tools/Hardware"),
        ("", "Accelerator", "Insta-Set™ CA Accelerator", "1 bottle", "Accelerates CA glue curing", "Ordered", "$8", "Included in Tools/Hardware"),
        
        # POWER SYSTEM
        ("POWER SYSTEM", "", "", "", "", "", "", ""),
        ("", "Battery", "OVONIC 4S LiPo Battery", "1-2", "14.8V nominal, 2200mAh, 130C, XT60", "Ordered", "$45", "https://a.co/d/1qcOXx2"),
        ("", "Charger", "B6 Battery Charger 80W", "1", "80W, LiPo Balance Charger/Discharger", "Ordered", "$30", "https://a.co/d/0vfvAfl"),
        ("", "Safety", "LiPo Safety Bag", "1-2", "Fire-resistant charging bag", "Ordered", "$10", "https://a.co/d/dm49smh"),
        ("", "Power Connectors", "XT60 Connectors (Male/Female Pairs)", "5 pairs", "High-current battery connectors", "Ordered", "$8", "https://a.co/d/XRBKSY2"),
        ("", "Power Distribution", "XT60 Y-Harness Splitter Cable", "1", "1 Female to 2 Male XT60", "NEEDED", "$8-12", "For dual-BEC power distribution"),
        ("", "Solar Panels", "Monocrystalline Mini Solar Cells 4W", "2", "4W each, Monocrystalline", "Ordered", "$24", "https://a.co/d/gdNsXPv"),
        ("", "Solar Controller", "DROK Boost Buck Converter", "1", "Solar charge controller/regulator", "Ordered", "$17", "https://a.co/d/0vfvAfl"),
        
        # MOTOR & PROPULSION
        ("MOTOR & PROPULSION", "", "", "", "", "", "", ""),
        ("", "Motor", "FLASH HOBBY D2830EVO Brushless Motor", "1", "1000KV, 2212 size, 14.8V rated", "Ordered", "$20", "https://a.co/d/1kmZM1H"),
        ("", "ESC", "40A Brushless ESC with 5V 3A BEC", "1", "40A continuous, 2-4S, Integrated BEC", "Ordered", "$13", "https://a.co/d/9lb3vMa"),
        ("", "Propeller", "Folding Propeller 1060/1160", "2 sets", "Folding design, with paddle clip adapter", "Ordered", "$12", "https://a.co/d/h3iRwcI"),
        ("", "Prop Adapter", "Propeller Adapter for 3.17mm Shaft", "1", "Aluminum adapter, collet-style", "NEEDED", "$8-10", "https://a.co/d/BZQWK83Z"),
        ("", "Motor Connectors", "3.5mm Bullet Connectors (Gold)", "1 set", "Male/Female, for motor 3-phase wires", "NEEDED", "$5-8", "Solder-type, 3-phase connection"),
        
        # FLIGHT CONTROL SYSTEM
        ("FLIGHT CONTROL SYSTEM", "", "", "", "", "", "", ""),
        ("", "Flight Controller", "MATEK F405-WMN Flight Controller", "1", "STM32F405, OSD, Blackbox, 2-6S", "Ordered", "$65", "ArduPilot/iNav compatible"),
        ("", "GPS Module", "BN-220 GPS Module TTL Dual GPS", "1", "UART, Compass, GPS+GLONASS", "Ordered", "$23", "https://a.co/d/5AzMOpc"),
        ("", "Radio TX", "Jumper Smart ELRS 2.4G RC Transmitter", "1", "ExpressLRS, 2.4GHz, Multi-protocol", "Ordered", "$65", "https://a.co/d/hAO1MYe"),
        ("", "Radio RX", "HAPPYMODEL ELRS EP2 Nano Receiver", "1", "2.4GHz, Long Range, UART/SBUS output", "Ordered", "$24", "https://a.co/d/hY74V3x"),
        ("", "BEC (External)", "Hobbywing 5V 6A UBEC (External)", "1", "5V @ 6A output, 2-6S input, Switching", "🚨 CRITICAL", "$12-15", "MUST ORDER - Servo power rail"),
        
        # SERVOS & CONTROL SURFACES
        ("SERVOS & CONTROL SURFACES", "", "", "", "", "", "", ""),
        ("", "Servos", "RGBZONE ES08MA II Metal Gear Servos", "4", "Metal gears, 9g, 180° rotation, 5V", "Ordered", "$24", "Ailerons (2), Elevator, Rudder"),
        ("", "Servo Extensions", "Servo Extension Cables 300mm", "4-6", "Male-Female, 22AWG, JR/Futaba std", "NEEDED", "$8-10", "https://a.co/d/C18BXDV"),
        ("", "Control Linkages", "Push Rods & Connector Kit", "1 kit", "Push rods, clevises, control horns", "NEEDED", "$10-12", "https://a.co/d/X9BWM6B"),
        
        # RESEARCH PAYLOAD
        ("RESEARCH PAYLOAD SYSTEM", "", "", "", "", "", "", ""),
        ("", "Microcontroller", "Arduino Nano V3.0 with Cable", "1-2", "ATmega328P, USB cable included", "Ordered", "$8", "https://a.co/d/e5A5Xpw"),
        ("", "Atmospheric Sensor", "HiLetgo BME280 Sensor Module", "1-2", "Temperature, Humidity, Pressure, I2C", "Ordered", "$9", "https://a.co/d/568Ea4O"),
        ("", "Gas Sensor", "Ximimark MQ135 Air Quality Sensor", "3", "Air quality, CO2, NH3, benzene", "Ordered", "$9", "https://a.co/d/7JlDKiE"),
        ("", "Data Storage", "SanDisk 32GB MicroSD Card", "2", "Class 10, High Speed, 2-pack", "Ordered", "$10", "https://a.co/d/9OyR5tD"),
        ("", "SD Card Reader", "HiLetgo MicroSD Card Adapter Module", "5", "SPI interface, 5V/3.3V compatible", "Ordered", "$11", "https://a.co/d/0a0SZVi"),
        ("", "Data Logger Switch", "SPST Mini Toggle Switch 2-Pin", "10", "Manual trigger for data logging", "Ordered", "$8", "https://a.co/d/f613hYO"),
        
        # WIRING & CONNECTORS
        ("WIRING & CONNECTORS", "", "", "", "", "", "", ""),
        ("", "Power Wire", "Silicone Wire 14AWG Red/Black", "2m ea", "High-temp silicone, flexible, 14AWG", "NEEDED", "$10-12", "For battery to ESC/UBEC (high current)"),
        ("", "Power Wire", "Silicone Wire 18AWG Red/Black", "3m ea", "High-temp silicone, 18AWG", "NEEDED", "$8-10", "For 5V power distribution"),
        ("", "Signal Wire", "Silicone Wire 22-24AWG (Multi-color)", "5m", "Red, Black, Yellow, White, Orange", "NEEDED", "$10-15", "Signal wires, UART, I2C, etc."),
        ("", "JST Connectors", "JST-SH 1.0mm Connector Kit", "1 kit", "4-pin for GPS, 2-pin for sensors", "NEEDED", "$8-10", "https://a.co/d/XRBKSY2"),
        ("", "Dupont Connectors", "Dupont 2.54mm Pin Connector Kit", "1 kit", "Male/Female headers, Arduino connections", "NEEDED", "$8-10", "For Arduino and sensor wiring"),
        ("", "Heat Shrink", "Heat Shrink Tubing Assortment", "1 kit", "Various sizes, 2:1 shrink ratio", "Ordered", "$10", "Included in Tools/Hardware"),
        
        # TOOLS & EQUIPMENT
        ("TOOLS & EQUIPMENT", "", "", "", "", "", "", ""),
        ("", "Soldering Station", "80W Adjustable Soldering Iron Kit", "1", "80W, Temperature control, w/ stand", "Ordered", "$24", "https://a.co/d/8bsmLsD"),
        ("", "Solder", "60/40 Rosin Core Solder", "1 roll", "0.8mm diameter, 60/40 tin/lead", "NEEDED", "$5-8", "Included in soldering kit or separate"),
        ("", "Wire Strippers", "Self-Adjusting Wire Stripper Tool", "1", "20-10 AWG capacity", "NEEDED", "$10-15", "For all wiring work"),
        ("", "Crimping Tool", "Dupont/JST Crimping Tool", "1", "PA-09/SN-28B style, ratcheting", "NEEDED", "$15-20", "For connector crimping"),
        ("", "Multimeter", "Digital Multimeter", "1", "Voltage, current, continuity testing", "NEEDED", "$15-25", "Essential for troubleshooting"),
        ("", "Hobby Knife Set", "X-Acto Style Hobby Knife with Blades", "1", "Precision cutting, trimming", "Ordered", "$8", "Included in Tools/Hardware"),
        ("", "CA Glue Debonder", "CA Glue Remover/Debonder", "1", "Removes cyanoacrylate mistakes", "NEEDED", "$8-10", "Helpful for build errors"),
        ("", "Zip Ties", "Cable Ties Assortment (100-200pcs)", "1 pkg", "Various sizes, 4-8 inch", "NEEDED", "$8-10", "Cable management and securing"),
        ("", "Double-Sided Tape", "3M VHB Double-Sided Tape", "1 roll", "Heavy-duty foam tape", "NEEDED", "$10-12", "Mounting electronics, sensors"),
        ("", "Velcro Straps", "Battery Velcro Straps", "2-4", "Hook & loop, 20mm x 300mm", "NEEDED", "$8-10", "Secure battery to airframe"),
        
        # OPTIONAL / FUTURE UPGRADES
        ("OPTIONAL / FUTURE UPGRADES", "", "", "", "", "", "", ""),
        ("", "OSD Module", "Standalone OSD Module", "1", "For FPV video overlay (future)", "Optional", "$15-20", "F405-WMN has integrated OSD"),
        ("", "Camera", "1000TVL FPV Camera", "1", "For FPV video feed (future upgrade)", "Optional", "$20-30", "Not required for initial flights"),
        ("", "VTX", "5.8GHz Video Transmitter", "1", "25-200mW VTX for FPV (future)", "Optional", "$20-30", "Not required for initial flights"),
        ("", "Antenna Tracker", "DIY Antenna Tracker Components", "1 set", "For long-range tracking (advanced)", "Optional", "$50+", "Future project consideration"),
        ("", "Pitot Tube", "Airspeed Sensor (Pitot Tube)", "1", "For true airspeed measurement", "Optional", "$15-25", "Advanced flight tuning"),
    ]
    
    # Status color mapping
    status_colors = {
        "Ordered": "C6EFCE",  # Light green
        "NEEDED": "FFEB9C",   # Light yellow
        "🚨 CRITICAL": "FFC7CE",  # Light red
        "Optional": "E7E6E6",  # Light gray
    }
    
    # Add parts data
    current_row = 5
    for part in parts_data:
        checkbox, category, component, qty, specs, status, price, notes = part
        
        # Check if this is a category header
        if category and not component:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            cell = ws.cell(row=current_row, column=1, value=category)
            cell.font = category_font
            cell.fill = category_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.row_dimensions[current_row].height = 25
        else:
            # Regular part row
            ws.cell(row=current_row, column=1, value=checkbox).alignment = center_align
            ws.cell(row=current_row, column=2, value=category).alignment = left_align
            ws.cell(row=current_row, column=3, value=component).alignment = left_align
            ws.cell(row=current_row, column=4, value=qty).alignment = center_align
            ws.cell(row=current_row, column=5, value=specs).alignment = left_align
            ws.cell(row=current_row, column=6, value=status).alignment = center_align
            ws.cell(row=current_row, column=7, value=price).alignment = center_align
            ws.cell(row=current_row, column=8, value=notes).alignment = left_align
            
            # Apply borders and fonts
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                cell.font = normal_font
            
            # Color code status column
            status_cell = ws.cell(row=current_row, column=6)
            if status in status_colors:
                status_cell.fill = PatternFill(start_color=status_colors[status], 
                                              end_color=status_colors[status], 
                                              fill_type="solid")
                if status == "🚨 CRITICAL":
                    status_cell.font = Font(name='Helvetica', size=10, bold=True, color="9C0006")
            
            # Make price column green and bold
            if price and price.startswith("$"):
                ws.cell(row=current_row, column=7).font = price_font
            
            # Make links blue and underlined
            if notes and ("http" in notes or "a.co" in notes):
                ws.cell(row=current_row, column=8).font = link_font
        
        current_row += 1
    
    # Add summary section
    current_row += 1
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws.cell(row=current_row, column=1, value="PROCUREMENT SUMMARY")
    cell.font = category_font
    cell.fill = category_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws.row_dimensions[current_row].height = 25
    
    current_row += 1
    summary_data = [
        ("Total Components Ordered:", "~40+ items", "Estimated Cost:", "$508"),
        ("Critical Items Needed:", "2 items", "UBEC + XT60 Y-Harness:", "~$20-27"),
        ("Recommended Items Needed:", "~15-20 items", "Wiring, Connectors, Tools:", "~$150-250"),
        ("Optional Upgrades:", "5+ items", "FPV, Sensors, Advanced:", "~$100+"),
        ("", "", "PROJECT TOTAL (Complete):", "$650-850"),
    ]
    
    for summary_row in summary_data:
        ws.cell(row=current_row, column=2, value=summary_row[0]).font = bold_font
        ws.cell(row=current_row, column=3, value=summary_row[1]).font = normal_font
        ws.cell(row=current_row, column=5, value=summary_row[2]).font = bold_font
        ws.cell(row=current_row, column=6, value=summary_row[3]).font = price_font
        
        for col in range(1, 9):
            ws.cell(row=current_row, column=col).border = thin_border
        
        current_row += 1
    
    # Add notes section
    current_row += 2
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws.cell(row=current_row, column=1, value="IMPORTANT NOTES")
    cell.font = category_font
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.font = Font(name='Helvetica', size=11, bold=True, color="FFFFFF")
    cell.alignment = center_align
    ws.row_dimensions[current_row].height = 25
    
    current_row += 1
    notes = [
        "🚨 CRITICAL: External 5V 6A UBEC must be ordered immediately - required for Dual-BEC architecture",
        "⚠ XT60 Y-Harness Splitter is required to split battery power to ESC and External UBEC",
        "⚠ Wiring supplies (14AWG, 18AWG, signal wire) are essential - cannot complete build without them",
        "⚠ Soldering equipment and tools are necessary for assembly - verify soldering kit includes solder",
        "✓ All major components (FC, GPS, Motor, ESC, Battery, Servos) have been ordered",
        "✓ Use this checklist during build to verify all parts are available before starting assembly",
        "📋 Check off items in the ✓ column as they arrive or are confirmed in stock",
        "💰 Budget tracking: Update prices as items are purchased to track actual vs. estimated costs",
    ]
    
    for note in notes:
        ws.cell(row=current_row, column=2, value=note).font = normal_font
        ws.merge_cells(f'B{current_row}:H{current_row}')
        ws.cell(row=current_row, column=2).alignment = left_align
        ws.row_dimensions[current_row].height = 20
        current_row += 1
    
    # Freeze panes (freeze header rows)
    ws.freeze_panes = 'A5'
    
    # Save workbook
    output_file = "UAV_Parts_Checklist.xlsx"
    wb.save(output_file)
    
    print(f"✓ Parts checklist created successfully!")
    print()
    print(f"📄 Output file: {output_file}")
    print()
    print("=" * 80)
    print("CHECKLIST SUMMARY:")
    print("=" * 80)
    print("✓ Organized by category (Airframe, Power, Flight Control, etc.)")
    print("✓ Includes specifications, quantities, and pricing")
    print("✓ Status color-coding (Green=Ordered, Yellow=Needed, Red=Critical)")
    print("✓ Amazon links included for reference")
    print("✓ Procurement summary with cost breakdown")
    print("✓ Ready for printing and use during build")
    print()
    print("🚨 CRITICAL ITEMS NEEDED:")
    print("   • External 5V 6A UBEC (Hobbywing recommended) - $12-15")
    print("   • XT60 Y-Harness Splitter - $8-12")
    print()
    print("⚠ RECOMMENDED ITEMS NEEDED:")
    print("   • Wiring supplies (14AWG, 18AWG, signal wire)")
    print("   • Connectors (JST, Dupont, bullet connectors)")
    print("   • Tools (wire strippers, crimpers, multimeter)")
    print()
    print("=" * 80)
    print()

if __name__ == "__main__":
    create_parts_checklist()
